from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from .models import Entry, AuditLog
from .forms import EntryForm, InterestCalculationForm, EntryFilterForm
from users.views import is_approved_user
from django.db.models import Q
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime

@login_required
@user_passes_test(is_approved_user)
def entry_create(request):
    if request.method == 'POST':
        form = EntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.user = request.user
            entry.save()
            AuditLog.objects.create(
                entry=entry,
                user=request.user,
                action='create',
                details='Entry created'
            )
            messages.success(request, 'Entry created successfully!')
            return redirect('entry_list')
    else:
        form = EntryForm()
    return render(request, 'entries/entry_form.html', {'form': form, 'action': 'Create'})

@login_required
@user_passes_test(is_approved_user)
def entry_list(request):
    entries = Entry.objects.filter(user=request.user, status='active')
    return render(request, 'entries/entry_list.html', {'entries': entries})

@login_required
@user_passes_test(is_approved_user)
def entry_edit(request, pk):
    entry = get_object_or_404(Entry, pk=pk, user=request.user, status='active')
    if request.method == 'POST':
        form = EntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                entry=entry,
                user=request.user,
                action='edit',
                details='Entry edited'
            )
            messages.success(request, 'Entry updated successfully!')
            return redirect('entry_list')
    else:
        form = EntryForm(instance=entry)
    return render(request, 'entries/entry_form.html', {'form': form, 'action': 'Edit'})

@login_required
@user_passes_test(is_approved_user)
def calculate_interest(request, pk):
    try:
        entry = Entry.objects.get(pk=pk)
        # Check if user is staff or the entry belongs to the user
        if not (request.user.is_staff or entry.user == request.user):
            messages.error(request, 'You do not have permission to access this entry.')
            return redirect('entry_list')
            
        calculation_result = None
        
        if request.method == 'POST':
            form = InterestCalculationForm(request.POST)
            if form.is_valid():
                daily_rate = form.cleaned_data['daily_rate']
                to_date = form.cleaned_data['to_date']
                
                if to_date:
                    entry.to_date = to_date
                else:
                    entry.to_date = timezone.now().date()
                    
                interest_amount = entry.calculate_interest(daily_rate)
                days = (entry.to_date - entry.from_date).days
                interest_type = "Compound" if days >= 365 else "Simple"
                
                calculation_result = {
                    'daily_rate': daily_rate,
                    'days': days,
                    'interest_type': interest_type,
                    'interest_amount': interest_amount,
                    'total_amount': entry.amount + interest_amount
                }
                
                entry.interest_rate = daily_rate
                entry.interest_amount = interest_amount
                entry.save()
                
                AuditLog.objects.create(
                    entry=entry,
                    user=request.user,
                    action='calculate_interest',
                    details=f'Interest calculated with rate {daily_rate}% for {days} days'
                )
                messages.success(request, 'Interest calculated successfully!')
        else:
            form = InterestCalculationForm()
            
        return render(request, 'entries/calculate_interest.html', {
            'form': form,
            'entry': entry,
            'calculation_result': calculation_result
        })
    except Entry.DoesNotExist:
        messages.error(request, 'Entry not found.')
        return redirect('entry_list')

@login_required
@user_passes_test(is_approved_user)
def release_entry(request, pk):
    entry = get_object_or_404(Entry, pk=pk, user=request.user, status='active')
    entry.release()
    AuditLog.objects.create(
        entry=entry,
        user=request.user,
        action='release',
        details='Entry released'
    )
    messages.success(request, 'Entry released successfully!')
    return redirect('entry_list')

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_dashboard(request):
    form = EntryFilterForm(request.GET)
    entries = Entry.objects.all()
    
    if form.is_valid():
        if form.cleaned_data['status']:
            entries = entries.filter(status=form.cleaned_data['status'])
        if form.cleaned_data['date_from']:
            entries = entries.filter(date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data['date_to']:
            entries = entries.filter(date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data['customer_name']:
            entries = entries.filter(customer_name__icontains=form.cleaned_data['customer_name'])
    
    stats = {
        'total_entries': entries.count(),
        'active_entries': entries.filter(status='active').count(),
        'released_entries': entries.filter(status='released').count(),
        'total_principal': entries.aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_interest': entries.aggregate(Sum('interest_amount'))['interest_amount__sum'] or 0,
    }
    
    audit_logs = AuditLog.objects.all().order_by('-timestamp')[:50]
    
    return render(request, 'entries/admin_dashboard.html', {
        'entries': entries,
        'stats': stats,
        'audit_logs': audit_logs,
        'form': form
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def released_entries(request):
    search_query = request.GET.get('search', '')
    entries = Entry.objects.filter(status='released')
    
    if search_query:
        entries = entries.filter(
            Q(serial_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    return render(request, 'entries/released_entries.html', {
        'entries': entries,
        'search_query': search_query
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_to_excel(request):
    # Get filter parameters
    status = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    entries = Entry.objects.all()
    
    # Apply filters
    if status != 'all':
        entries = entries.filter(status=status)
    if date_from:
        entries = entries.filter(date__gte=date_from)
    if date_to:
        entries = entries.filter(date__lte=date_to)
    if search_query:
        entries = entries.filter(
            Q(serial_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries Report"
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write headers
    headers = [
        'Date', 'User', 'Serial Number', 'Customer Name', 'Amount', 
        'Weight', 'Status', 'Interest Rate', 'Interest Amount', 
        'Total Amount', 'Created At', 'Updated At'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row, entry in enumerate(entries, 2):
        # Convert timezone-aware datetimes to naive datetimes
        created_at = entry.created_at.replace(tzinfo=None) if entry.created_at else None
        updated_at = entry.updated_at.replace(tzinfo=None) if entry.updated_at else None
        
        ws.cell(row=row, column=1, value=entry.date)
        ws.cell(row=row, column=2, value=entry.user.username)
        ws.cell(row=row, column=3, value=entry.serial_number)
        ws.cell(row=row, column=4, value=entry.customer_name)
        ws.cell(row=row, column=5, value=float(entry.amount))
        ws.cell(row=row, column=6, value=float(entry.weight))
        ws.cell(row=row, column=7, value=entry.status)
        ws.cell(row=row, column=8, value=float(entry.interest_rate) if entry.interest_rate else None)
        ws.cell(row=row, column=9, value=float(entry.interest_amount) if entry.interest_amount else None)
        ws.cell(row=row, column=10, value=float(entry.amount + entry.interest_amount) if entry.interest_amount else float(entry.amount))
        ws.cell(row=row, column=11, value=created_at)
        ws.cell(row=row, column=12, value=updated_at)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=entries_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response
